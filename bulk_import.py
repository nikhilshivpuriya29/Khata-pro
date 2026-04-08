import csv
import io
from datetime import datetime, date
from flask import Blueprint, render_template, request, flash, redirect, url_for, Response
from flask_login import login_required, current_user
from models import db, Customer, Transaction

import_bp = Blueprint("bulk", __name__)

# CSV column headers
CUSTOMER_HEADERS = [
    "full_name", "father_name", "phone", "aadhaar",
    "address_street", "address_city", "address_district", "address_state", "address_pin",
]

TRANSACTION_HEADERS = [
    "customer_phone", "txn_type", "amount", "item_description",
    "purchase_date", "promised_date", "note",
]


def parse_date(val):
    """Try common date formats."""
    if not val:
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(val.strip(), fmt).date()
        except ValueError:
            continue
    return None


@import_bp.route("/import", methods=["GET"])
@login_required
def import_page():
    return render_template("import.html")


@import_bp.route("/import/template/customers")
@login_required
def download_customer_template():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(CUSTOMER_HEADERS)
    # Sample row
    writer.writerow([
        "Ramesh Kumar", "Suresh Kumar", "9876543210", "123456789012",
        "123 Main Bazaar", "Jaipur", "Jaipur", "Rajasthan", "302001",
    ])
    output.seek(0)
    return Response(
        output.getvalue(), mimetype="text/csv",
        headers={"Content-Disposition": "attachment;filename=customer_template.csv"},
    )


@import_bp.route("/import/template/transactions")
@login_required
def download_txn_template():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(TRANSACTION_HEADERS)
    writer.writerow([
        "9876543210", "credit", "5000", "10 bags cement",
        "2026-01-15", "2026-02-15", "Monthly supply",
    ])
    writer.writerow([
        "9876543210", "debit", "2000", "",
        "2026-02-01", "", "Partial payment",
    ])
    output.seek(0)
    return Response(
        output.getvalue(), mimetype="text/csv",
        headers={"Content-Disposition": "attachment;filename=transaction_template.csv"},
    )


def read_file_rows(file):
    """Read CSV or Excel file and return list of dicts."""
    filename = file.filename.lower()
    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        from openpyxl import load_workbook
        wb = load_workbook(file, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip().lower() if h else "" for h in rows[0]]
        data = []
        for row in rows[1:]:
            d = {}
            for i, h in enumerate(headers):
                val = row[i] if i < len(row) else None
                d[h] = str(val).strip() if val is not None else ""
            data.append(d)
        return data
    else:
        # CSV
        content = file.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(content))
        return [{k.strip().lower(): (v.strip() if v else "") for k, v in row.items()} for row in reader]


@import_bp.route("/import/customers", methods=["POST"])
@login_required
def import_customers():
    file = request.files.get("file")
    if not file or not file.filename:
        flash("Please select a file.", "danger")
        return redirect(url_for("bulk.import_page"))

    try:
        rows = read_file_rows(file)
    except Exception as e:
        flash(f"Error reading file: {e}", "danger")
        return redirect(url_for("bulk.import_page"))

    added = 0
    skipped = 0
    errors = []

    for i, row in enumerate(rows, start=2):
        name = row.get("full_name", "").strip()
        phone = row.get("phone", "").strip()
        address_street = row.get("address_street", "").strip()
        address_city = row.get("address_city", "").strip()

        # Mandatory: name, phone, at least one address field
        if not name:
            errors.append(f"Row {i}: Missing full_name")
            skipped += 1
            continue
        if not phone:
            errors.append(f"Row {i}: Missing phone for '{name}'")
            skipped += 1
            continue
        if not address_street and not address_city:
            errors.append(f"Row {i}: Missing address for '{name}'")
            skipped += 1
            continue

        # Check duplicate by phone
        existing = Customer.query.filter_by(user_id=current_user.id, phone=phone).first()
        if existing:
            errors.append(f"Row {i}: Phone '{phone}' already exists ({existing.full_name})")
            skipped += 1
            continue

        c = Customer(
            user_id=current_user.id,
            full_name=name,
            father_name=row.get("father_name", "").strip(),
            phone=phone,
            aadhaar=row.get("aadhaar", "").strip(),
            address_street=address_street,
            address_city=address_city,
            address_district=row.get("address_district", "").strip(),
            address_state=row.get("address_state", "").strip(),
            address_pin=row.get("address_pin", "").strip(),
        )
        db.session.add(c)
        added += 1

    db.session.commit()
    msg = f"Imported {added} customers."
    if skipped:
        msg += f" Skipped {skipped}."
    flash(msg, "success" if added > 0 else "warning")

    if errors:
        for err in errors[:10]:
            flash(err, "warning")
        if len(errors) > 10:
            flash(f"...and {len(errors) - 10} more errors.", "warning")

    return redirect(url_for("bulk.import_page"))


@import_bp.route("/import/transactions", methods=["POST"])
@login_required
def import_transactions():
    file = request.files.get("file")
    if not file or not file.filename:
        flash("Please select a file.", "danger")
        return redirect(url_for("bulk.import_page"))

    try:
        rows = read_file_rows(file)
    except Exception as e:
        flash(f"Error reading file: {e}", "danger")
        return redirect(url_for("bulk.import_page"))

    added = 0
    skipped = 0
    errors = []

    for i, row in enumerate(rows, start=2):
        phone = row.get("customer_phone", "").strip()
        txn_type = row.get("txn_type", "").strip().lower()
        amount_str = row.get("amount", "").strip()

        if not phone:
            errors.append(f"Row {i}: Missing customer_phone")
            skipped += 1
            continue
        if txn_type not in ("credit", "debit"):
            errors.append(f"Row {i}: Invalid txn_type '{txn_type}' (use credit/debit)")
            skipped += 1
            continue
        try:
            amount = float(amount_str)
            if amount <= 0:
                raise ValueError
        except (ValueError, TypeError):
            errors.append(f"Row {i}: Invalid amount '{amount_str}'")
            skipped += 1
            continue

        customer = Customer.query.filter_by(user_id=current_user.id, phone=phone).first()
        if not customer:
            errors.append(f"Row {i}: No customer with phone '{phone}'")
            skipped += 1
            continue

        purchase_date = parse_date(row.get("purchase_date", "")) or date.today()
        promised_date = parse_date(row.get("promised_date", ""))

        txn = Transaction(
            customer_id=customer.id,
            txn_type=txn_type,
            amount=amount,
            item_description=row.get("item_description", "").strip(),
            purchase_date=purchase_date,
            promised_date=promised_date,
            note=row.get("note", "").strip(),
        )
        db.session.add(txn)
        added += 1

    db.session.commit()
    msg = f"Imported {added} transactions."
    if skipped:
        msg += f" Skipped {skipped}."
    flash(msg, "success" if added > 0 else "warning")

    if errors:
        for err in errors[:10]:
            flash(err, "warning")
        if len(errors) > 10:
            flash(f"...and {len(errors) - 10} more errors.", "warning")

    return redirect(url_for("bulk.import_page"))
